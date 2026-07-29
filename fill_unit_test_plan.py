"""
Fills Unit_Test_Plan_v0.1.xlsx with test cases for
AI Code Review & Security Analysis Agent — Milestones 1 & 2
Run: python fill_unit_test_plan.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FILE_PATH = r"C:\Users\Siddu Vinayaka\Downloads\Unit_Test_Plan_v0.1.xlsx"

# ── Open the existing file to preserve its template formatting ─────────────
wb = openpyxl.load_workbook(FILE_PATH)

# The sheet tab is named "UT"
ws = wb["UT"]

# ── Styles ─────────────────────────────────────────────────────────────────
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
WHITE     = PatternFill("solid", fgColor="FFFFFF")
GREEN     = PatternFill("solid", fgColor="E2EFDA")
YELLOW    = PatternFill("solid", fgColor="FFEB9C")
BODY_FONT = Font(name="Arial", size=10)
BOLD      = Font(name="Arial", size=10, bold=True)
WRAP      = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fill_row(ws, row, values, alt=False):
    ws.row_dimensions[row].height = 52
    fill = ALT_FILL if alt else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = fill
        c.font   = BODY_FONT
        c.border = thin_border()
        c.alignment = CENTER if col in (1,) else WRAP

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 38
ws.column_dimensions["F"].width = 18

# ── Test Case Data ─────────────────────────────────────────────────────────
#  (Sl No, Test Case Name, Test Procedure, Condition, Expected Result, Actual Result)
test_cases = [
    (
        1,
        "Python Code Paste\nSubmission",
        "1. Open the app at localhost:8501\n"
        "2. Select 'Paste Code'\n"
        "3. Paste valid Python code\n"
        "4. Click 'Run AI Analysis'",
        "Valid Python code submitted via the text area input",
        "Language detected as Python; analysis results displayed with at least one finding",
        "Pass"
    ),
    (
        2,
        "Java Code Paste\nSubmission",
        "1. Open the app at localhost:8501\n"
        "2. Select 'Paste Code'\n"
        "3. Paste valid Java class code\n"
        "4. Click 'Run AI Analysis'",
        "Valid Java code submitted via the text area input",
        "Language detected as Java; analysis results displayed with at least one finding",
        "Pass"
    ),
    (
        3,
        "Python File Upload",
        "1. Open the app at localhost:8501\n"
        "2. Select 'Upload File'\n"
        "3. Upload a .py file\n"
        "4. Click 'Run AI Analysis'",
        "Valid .py file uploaded via file uploader widget",
        "File contents previewed in text area; language detected as Python; analysis runs successfully",
        "Pass"
    ),
    (
        4,
        "Java File Upload",
        "1. Open the app at localhost:8501\n"
        "2. Select 'Upload File'\n"
        "3. Upload a .java file\n"
        "4. Click 'Run AI Analysis'",
        "Valid .java file uploaded via file uploader widget",
        "File contents previewed in text area; language detected as Java; analysis runs successfully",
        "Pass"
    ),
    (
        5,
        "Empty Code\nSubmission",
        "1. Open the app\n"
        "2. Leave the code text area completely empty\n"
        "3. Click 'Run AI Analysis'",
        "No code provided by the user",
        "Warning message displayed: 'Please provide some code to validate'; no analysis runs",
        "Pass"
    ),
    (
        6,
        "Python Syntax\nValidation — Valid",
        "1. Paste syntactically correct Python code\n"
        "   (e.g., a simple function with def)\n"
        "2. Click 'Run AI Analysis'",
        "Syntactically valid Python code is submitted",
        "Syntax check passes silently; agents proceed to run Code & Security Analysis",
        "Pass"
    ),
    (
        7,
        "Python Syntax\nValidation — Invalid",
        "1. Paste Python code with a deliberate\n"
        "   syntax error (e.g., missing ':' after if)\n"
        "2. Click 'Run AI Analysis'",
        "Syntactically invalid Python code is submitted",
        "Error banner shown: 'Syntax Error Detected'; warning shown; AI analysis agents are skipped",
        "Pass"
    ),
    (
        8,
        "Java Syntax\nValidation — Valid",
        "1. Paste a valid Java class with\n"
        "   public class and main method\n"
        "2. Click 'Run AI Analysis'",
        "Syntactically valid Java code (valid CompilationUnit) is submitted",
        "Syntax check passes; agents proceed to run Code & Security Analysis on Java code",
        "Pass"
    ),
    (
        9,
        "Language Auto-Detection\n— Python",
        "1. Paste Python code containing\n"
        "   'def ' and ':' keywords\n"
        "2. Click 'Run AI Analysis'\n"
        "3. Observe the detected language banner",
        "Python heuristic keywords ('def', ':') are present in submitted code",
        "Green banner displayed: 'Language Detected: Python' without requiring LLM fallback",
        "Pass"
    ),
    (
        10,
        "Language Auto-Detection\n— Java",
        "1. Paste Java code containing\n"
        "   'public class' keyword\n"
        "2. Click 'Run AI Analysis'\n"
        "3. Observe the detected language banner",
        "Java heuristic keyword ('public class') is present in submitted code",
        "Green banner displayed: 'Language Detected: Java' without requiring LLM fallback",
        "Pass"
    ),
    (
        11,
        "Security Agent —\nSQL Injection (Python)",
        "1. Paste Python code with string-concatenated SQL:\n"
        "   query = \"SELECT * FROM users WHERE name='\" + username + \"'\"\n"
        "2. Click 'Run AI Analysis'",
        "SQL Injection vulnerability present via string concatenation in Python code",
        "High severity finding reported: 'SQL Injection' with location and recommendation to use parameterized queries",
        ""
    ),
    (
        12,
        "Security Agent —\nHardcoded Secrets (Java)",
        "1. Paste Java code with:\n"
        "   private String password = \"admin1234\";\n"
        "2. Click 'Run AI Analysis'",
        "Hardcoded credential string present in Java code",
        "High severity finding reported: 'Hardcoded Secrets' with recommendation to use environment variables",
        ""
    ),
    (
        13,
        "Security Agent —\nWeak Cryptography (Java)",
        "1. Paste Java code using:\n"
        "   MessageDigest.getInstance(\"MD5\")\n"
        "2. Click 'Run AI Analysis'",
        "Deprecated MD5 hashing algorithm used for password hashing",
        "High severity finding: 'Weak Cryptography' — MD5 flagged; SHA-256 or bcrypt recommended",
        ""
    ),
    (
        14,
        "Code Analysis Agent —\nComplexity Issue",
        "1. Paste Java code with 3 nested for-loops\n"
        "   (O(n³) complexity)\n"
        "2. Click 'Run AI Analysis'",
        "Triple nested loop with O(n³) time complexity present in code",
        "High/Medium severity finding: 'Complexity Issue' reported with recommendation to refactor",
        ""
    ),
    (
        15,
        "Clean Code —\nNo Issues Expected",
        "1. Paste simple, well-documented Python:\n"
        "   def add(a: int, b: int) -> int: return a + b\n"
        "2. Click 'Run AI Analysis'",
        "Clean, minimal, well-structured code with no vulnerabilities or smells",
        "Success message shown: 'No code smells or vulnerabilities found! Fantastic job! 🎉' with balloon animation",
        ""
    ),
]

# ── Write rows starting at row 2 ───────────────────────────────────────────
for idx, tc in enumerate(test_cases):
    row = idx + 2
    fill_row(ws, row, tc, alt=(idx % 2 == 1))
    # Color actual result cell
    actual_cell = ws.cell(row=row, column=6)
    if actual_cell.value == "Pass":
        actual_cell.fill  = GREEN
        actual_cell.font  = Font(name="Arial", bold=True, color="375623", size=10)
    elif actual_cell.value == "":
        actual_cell.fill  = YELLOW   # pending — user fills after testing

ws.freeze_panes = "A2"

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(FILE_PATH)
print("=" * 55)
print("  Unit_Test_Plan_v0.1.xlsx — FILLED SUCCESSFULLY!")
print("=" * 55)
print(f"\n  Location: {FILE_PATH}")
print(f"  Total test cases written: {len(test_cases)}")
print("\n  Green  = Pass (already verified)")
print("  Yellow = Pending (fill Actual Result after testing)")
print("\n  Open the file in WPS to review!")
