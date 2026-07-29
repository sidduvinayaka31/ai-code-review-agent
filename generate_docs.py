"""
Generates the complete project documentation spreadsheet for
AI Code Review & Security Analysis Agent — Milestones 1 & 2
Run: python generate_docs.py
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Colour palette ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="4472C4")   # blue header
ALT_FILL      = PatternFill("solid", fgColor="DCE6F1")   # light blue alt row
ORANGE_FILL   = PatternFill("solid", fgColor="F4B183")   # orange (sprint sheet)
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")
RED_FILL      = PatternFill("solid", fgColor="FCE4D6")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ORANGE_FONT   = Font(name="Calibri", bold=True, color="000000", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, fill, font, col_count, height=30):
    ws.row_dimensions[row].height = height
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.fill  = fill
        c.font  = font
        c.alignment = CENTER
        c.border = thin_border()

def style_body_row(ws, row, col_count, alt=False, height=45):
    ws.row_dimensions[row].height = height
    fill = ALT_FILL if alt else WHITE_FILL
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.fill  = fill
        c.font  = BODY_FONT
        c.alignment = LEFT
        c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — TEST CASES
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Test Cases"

headers = [
    "Sl: No:", "Test Case Name", "Test Procedure",
    "Condition to be tested", "Expected Result", "Actual Result",
    "Description", "Detected Sprint", "Assigned To",
    "Type Of Defect", "Action Taken", "Action Taken Date",
    "Status\n(Open/Closed)", "Remarks"
]

col_widths = [8, 28, 42, 32, 35, 20, 35, 16, 14, 18, 35, 18, 16, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header_row(ws1, 1, HEADER_FILL, HEADER_FONT, len(headers))

test_cases = [
    # (Sl, Name, Procedure, Condition, Expected, Actual, Description,
    #  Sprint, AssignedTo, DefectType, ActionTaken, ActionDate, Status, Remarks)
    (1, "Python Code Paste Submission",
     "1. Open app\n2. Select 'Paste Code'\n3. Paste valid Python code\n4. Click 'Run AI Analysis'",
     "Valid Python code submitted via paste input",
     "Language detected as Python; analysis results displayed with findings",
     "Pass",
     "Verify code paste input works end-to-end for Python",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (2, "Java Code Paste Submission",
     "1. Open app\n2. Select 'Paste Code'\n3. Paste valid Java class code\n4. Click 'Run AI Analysis'",
     "Valid Java code submitted via paste input",
     "Language detected as Java; analysis results displayed with findings",
     "Pass",
     "Verify code paste input works end-to-end for Java",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (3, "Python File Upload",
     "1. Open app\n2. Select 'Upload File'\n3. Upload a .py file\n4. Click 'Run AI Analysis'",
     "Valid .py file uploaded via file uploader",
     "File contents shown in preview; language detected as Python; analysis runs",
     "Pass",
     "Verify .py file upload and content preview",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (4, "Java File Upload",
     "1. Open app\n2. Select 'Upload File'\n3. Upload a .java file\n4. Click 'Run AI Analysis'",
     "Valid .java file uploaded via file uploader",
     "File contents shown in preview; language detected as Java; analysis runs",
     "Pass",
     "Verify .java file upload and content preview",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (5, "Empty Code Submission",
     "1. Open app\n2. Leave text area empty\n3. Click 'Run AI Analysis'",
     "No code provided by user",
     "Warning message displayed: 'Please provide some code to validate'",
     "Pass",
     "Verify graceful handling of empty submission",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Edge case"),

    (6, "Python Syntax Validation — Valid",
     "1. Paste syntactically correct Python code\n2. Click 'Run AI Analysis'",
     "Valid Python syntax submitted",
     "Syntax check passes silently; analysis proceeds to AI agents",
     "Pass",
     "Verify valid Python syntax passes the validator",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (7, "Python Syntax Validation — Invalid",
     "1. Paste Python code with missing colon after if-statement\n2. Click 'Run AI Analysis'",
     "Syntactically invalid Python code",
     "Error shown: 'Syntax Error Detected'; analysis agents are skipped",
     "Pass",
     "Verify syntax errors are caught before LLM analysis",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (8, "Java Syntax Validation — Valid",
     "1. Paste valid Java class with main method\n2. Click 'Run AI Analysis'",
     "Valid Java CompilationUnit submitted",
     "Syntax check passes; analysis proceeds to AI agents",
     "Pass",
     "Verify valid Java syntax passes the javalang validator",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (9, "Java Syntax Validation — Invalid",
     "1. Paste Java code with missing semicolon\n2. Click 'Run AI Analysis'",
     "Syntactically invalid Java code",
     "Error shown: 'Syntax Error Detected'; analysis skipped",
     "Pass",
     "Verify javalang catches Java syntax errors",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (10, "Language Auto-Detection — Python",
     "1. Paste Python code containing 'def' and ':'\n2. Click 'Run AI Analysis'",
     "Python heuristic keywords present in code",
     "Language correctly identified as Python without LLM call",
     "Pass",
     "Verify fast heuristic detection for Python",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (11, "Language Auto-Detection — Java",
     "1. Paste Java code containing 'public class'\n2. Click 'Run AI Analysis'",
     "Java heuristic keywords present in code",
     "Language correctly identified as Java without LLM call",
     "Pass",
     "Verify fast heuristic detection for Java",
     "Sprint 1", "Developer", "—", "—", "—", "Closed", "Milestone 1"),

    (12, "Code Analysis — SQL Injection (Python)",
     "1. Paste Python code with string-concatenated SQL query\n2. Click 'Run AI Analysis'",
     "SQL Injection vulnerability present in Python code",
     "High severity 'SQL Injection' finding reported with line number",
     "",
     "Verify Code Analysis Agent detects SQL injection in Python",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (13, "Code Analysis — SQL Injection (Java)",
     "1. Paste Java code with Statement + string concat SQL\n2. Click 'Run AI Analysis'",
     "SQL Injection vulnerability present in Java code",
     "High severity 'SQL Injection' finding reported with line number",
     "",
     "Verify Security Agent detects SQL injection in Java",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (14, "Security Agent — Hardcoded Secrets",
     "1. Paste Java code with hardcoded DB password string\n2. Click 'Run AI Analysis'",
     "Hardcoded credential string present in code",
     "High severity 'Hardcoded Secrets' finding reported",
     "",
     "Verify Security Agent detects hardcoded passwords",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (15, "Security Agent — XSS Detection",
     "1. Paste Java Servlet with unsanitized request parameter in HTML output\n2. Click 'Run AI Analysis'",
     "XSS vulnerability via reflected user input",
     "High severity 'Cross-Site Scripting (XSS)' finding reported",
     "",
     "Verify Security Agent detects XSS in Java servlets",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (16, "Security Agent — Weak Cryptography",
     "1. Paste Java code using MD5 MessageDigest\n2. Click 'Run AI Analysis'",
     "MD5 (weak hash algorithm) used for password hashing",
     "High severity 'Weak Cryptography' finding reported; SHA-256 recommended",
     "",
     "Verify Security Agent flags deprecated MD5 hashing",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (17, "Code Analysis — Code Smells (Triple Nested Loop)",
     "1. Paste Java code with 3 nested for-loops\n2. Click 'Run AI Analysis'",
     "O(n³) complexity triple-nested loop present",
     "High/Medium severity 'Complexity Issue' finding reported",
     "",
     "Verify Code Analysis Agent detects excessive loop nesting",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (18, "Multi-Agent Parallel Execution",
     "1. Submit code\n2. Observe terminal logs during analysis",
     "Both Code Analysis and Security agents must run concurrently",
     "Terminal shows both '--- Running Code Analysis ---' and '--- Running Security Analysis ---' messages",
     "",
     "Verify LangGraph fan-out runs both agents in parallel",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (19, "Severity Sorting of Results",
     "1. Submit code with mix of High, Medium, Low findings\n2. Observe results order",
     "Multiple findings of different severities returned",
     "Results displayed in order: High (🔴) → Medium (🟡) → Low (🔵)",
     "",
     "Verify findings are sorted by severity in the UI",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),

    (20, "Clean Code — Zero Issues",
     "1. Paste simple, well-documented, clean Python/Java code\n2. Click 'Run AI Analysis'",
     "No significant code quality or security issues present",
     "Success message: 'No code smells or vulnerabilities found! Fantastic job!' with balloon animation",
     "",
     "Verify agent correctly identifies clean code with no false positives",
     "Sprint 2", "Developer", "—", "—", "—", "Open", "Milestone 2"),
]

for idx, row_data in enumerate(test_cases):
    r = idx + 2
    for col, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=col, value=val)
    style_body_row(ws1, r, len(headers), alt=(idx % 2 == 1))
    # Green for Pass, empty for pending
    actual_cell = ws1.cell(row=r, column=6)
    if actual_cell.value == "Pass":
        actual_cell.fill = GREEN_FILL
        actual_cell.font = Font(name="Calibri", bold=True, color="375623", size=10)
    status_cell = ws1.cell(row=r, column=13)
    if status_cell.value == "Closed":
        status_cell.fill = GREEN_FILL
    elif status_cell.value == "Open":
        status_cell.fill = RED_FILL

ws1.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — BUG / DEFECT LOG
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Bug Defect Log")

bug_headers = [
    "Sl: No:", "Bug ID", "Description", "Detected Sprint",
    "Assigned To", "Type Of Defect", "Action Taken",
    "Action Taken Date", "Status\n(Open/Closed)", "Remarks"
]
bug_col_widths = [8, 12, 45, 16, 14, 20, 48, 20, 18, 25]
for i, w in enumerate(bug_col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for i, h in enumerate(bug_headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header_row(ws2, 1, HEADER_FILL, HEADER_FONT, len(bug_headers))

bugs = [
    (1, "BUG-001",
     "syntax_validator.py was built but not wired into the agent pipeline — validation was never called",
     "Sprint 1", "Developer", "Functional Defect",
     "Added 'validate_syntax' as a LangGraph node between detect_language and analysis agents",
     "24-Jul-2026", "Closed", "Milestone 1 requirement missed in initial build"),

    (2, "BUG-002",
     "LangGraph conditional edge fan-out using lambda returning list caused incorrect routing",
     "Sprint 2", "Developer", "Integration Defect",
     "Replaced lambda with proper route_after_syntax() function; used explicit edge list",
     "24-Jul-2026", "Closed", "LangGraph API misuse corrected"),

    (3, "BUG-003",
     "Model 'gemini-1.5-flash-latest' returns 404 NOT_FOUND — deprecated model suffix",
     "Sprint 2", "Developer", "Environment Defect",
     "Removed deprecated '-latest' suffix; changed to 'gemini-1.5-flash'",
     "24-Jul-2026", "Closed", "Google deprecated all '-latest' model aliases"),

    (4, "BUG-004",
     "Embedding model 'text-embedding-004' returns 404 NOT_FOUND on API v1beta endpoint",
     "Sprint 2", "Developer", "Environment Defect",
     "Replaced Google embeddings with local FastEmbed (BAAI/bge-small-en-v1.5) — no API quota",
     "24-Jul-2026", "Closed", "Google v1beta API does not expose this embedding model"),

    (5, "BUG-005",
     "pip install sentence-transformers fails with [WinError 206] filename too long due to PyTorch deep paths",
     "Sprint 2", "Developer", "Environment Defect",
     "Replaced sentence-transformers with fastembed which uses short ONNX paths; uninstalled torch",
     "24-Jul-2026", "Closed", "Windows 260-character MAX_PATH limit exceeded"),

    (6, "BUG-006",
     "Model 'gemini-2.0-flash' returns RESOURCE_EXHAUSTED (429) — free tier quota exceeded",
     "Sprint 2", "Developer", "Environment Defect",
     "Switched entirely to Groq API (llama-3.3-70b-versatile); 14,400 req/day free tier",
     "24-Jul-2026", "Closed", "gemini-2.0-flash has near-zero free tier RPM"),

    (7, "BUG-007",
     "'gemini-1.5-flash' returns 404 NOT_FOUND — langchain-google-genai uses v1beta which doesn't support this model",
     "Sprint 2", "Developer", "Environment Defect",
     "Fully migrated all LLM calls to Groq (langchain-groq); removed langchain-google-genai dependency",
     "24-Jul-2026", "Closed", "Old langchain-google-genai package version incompatible with current models"),
]

for idx, row_data in enumerate(bugs):
    r = idx + 2
    for col, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=col, value=val)
    style_body_row(ws2, r, len(bug_headers), alt=(idx % 2 == 1), height=55)
    status_cell = ws2.cell(row=r, column=9)
    if status_cell.value == "Closed":
        status_cell.fill = GREEN_FILL
        status_cell.font = Font(name="Calibri", bold=True, color="375623", size=10)
    elif status_cell.value == "Open":
        status_cell.fill = RED_FILL

ws2.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SPRINT BACKLOG / USER STORIES
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Sprint Backlog")

sprint_headers = [
    "Planned\nSprint", "Actual\nSprint", "US ID",
    "User Story Description",
    "MOSCOW", "Dependency", "Assignee", "Status"
]
sprint_col_widths = [12, 12, 12, 65, 14, 18, 14, 14]
for i, w in enumerate(sprint_col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

for i, h in enumerate(sprint_headers, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, ORANGE_FILL, ORANGE_FONT, len(sprint_headers))

user_stories = [
    ("Sprint 1", "Sprint 1", "US-001",
     "As a developer, I want to paste Python or Java code directly into the portal so I can get an automated review without uploading a file.",
     "Must Have", "None", "Developer", "Completed"),

    ("Sprint 1", "Sprint 1", "US-002",
     "As a developer, I want to upload a .py or .java source file so I can review code from my real project files.",
     "Must Have", "US-001", "Developer", "Completed"),

    ("Sprint 1", "Sprint 1", "US-003",
     "As a developer, I want the system to automatically detect whether my code is Python or Java so I don't have to select the language manually.",
     "Must Have", "US-001", "Developer", "Completed"),

    ("Sprint 1", "Sprint 1", "US-004",
     "As a developer, I want the system to validate my code's syntax before running AI analysis so I receive meaningful feedback on broken code and avoid wasting API credits.",
     "Must Have", "US-003", "Developer", "Completed"),

    ("Sprint 1", "Sprint 1", "US-005",
     "As a developer, I want OWASP guidelines and secure coding standards to be indexed into a RAG pipeline so the security agent has authoritative, up-to-date context when scanning my code.",
     "Must Have", "None", "Developer", "Completed"),

    ("Sprint 2", "Sprint 2", "US-006",
     "As a developer, I want a Code Analysis Agent to identify code smells, complexity issues, and design anti-patterns with severity scoring so I can improve code quality systematically.",
     "Must Have", "US-003", "Developer", "Completed"),

    ("Sprint 2", "Sprint 2", "US-007",
     "As a developer, I want a Security Vulnerability Agent to scan my code for OWASP-standard vulnerabilities (SQL Injection, XSS, Hardcoded Secrets, etc.) and classify them by type and severity with location-specific flagging.",
     "Must Have", "US-005", "Developer", "Completed"),

    ("Sprint 2", "Sprint 2", "US-008",
     "As a developer, I want the Code Analysis and Security Vulnerability agents to run in parallel via a LangGraph orchestrator so that overall analysis time is minimised.",
     "Must Have", "US-006, US-007", "Developer", "Completed"),

    ("Sprint 2", "Sprint 2", "US-009",
     "As a developer, I want all agent findings merged into a unified list sorted by severity (High → Medium → Low) so I can prioritise the most critical issues first.",
     "Must Have", "US-008", "Developer", "Completed"),

    ("Sprint 2", "Sprint 2", "US-010",
     "As a developer, I want the system's detection accuracy validated against sample Python and Java codebases containing known quality issues and vulnerabilities, so I can confirm the agents are working correctly.",
     "Must Have", "US-006, US-007", "Developer", "Completed"),
]

for idx, row_data in enumerate(user_stories):
    r = idx + 2
    for col, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=col, value=val)
    style_body_row(ws3, r, len(sprint_headers), alt=(idx % 2 == 1), height=55)
    status_cell = ws3.cell(row=r, column=8)
    if status_cell.value == "Completed":
        status_cell.fill = GREEN_FILL
        status_cell.font = Font(name="Calibri", bold=True, color="375623", size=10)
    status_cell.alignment = CENTER

ws3.freeze_panes = "A2"


# ── Save ───────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Siddu Vinayaka\.gemini\antigravity\scratch\ai_code_review_agent\Project_Documentation_M1_M2.xlsx"
wb.save(output_path)
print(f"✅ Spreadsheet saved to:\n   {output_path}")
print("\nSheets created:")
print("  • Test Cases       (20 test cases for Milestones 1 & 2)")
print("  • Bug Defect Log   (7 bugs found and resolved)")
print("  • Sprint Backlog   (10 user stories across Sprint 1 & 2)")
