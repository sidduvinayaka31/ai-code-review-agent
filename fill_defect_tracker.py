"""
Fills Defect_Tracker Template_v0.1.xlsx with all bugs found and resolved
during AI Code Review & Security Analysis Agent — Milestones 1 & 2
Run: python fill_defect_tracker.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FILE_PATH = r"C:\Users\Siddu Vinayaka\Downloads\Defect_Tracker Template_v0.1.xlsx"

# ── Open existing file to preserve template formatting ─────────────────────
wb = openpyxl.load_workbook(FILE_PATH)

# Auto-detect the sheet name
print(f"Sheets found: {wb.sheetnames}")
ws = wb.active   # use the first/active sheet
print(f"Writing to sheet: '{ws.title}'")

# ── Styles ─────────────────────────────────────────────────────────────────
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
WHITE       = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
RED_FILL    = PatternFill("solid", fgColor="FCE4D6")
BODY_FONT   = Font(name="Arial", size=10)
GREEN_FONT  = Font(name="Arial", bold=True, color="375623", size=10)
RED_FONT    = Font(name="Arial", bold=True, color="9C0006", size=10)
WRAP        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fill_row(ws, row, values, alt=False, height=60):
    ws.row_dimensions[row].height = height
    fill = ALT_FILL if alt else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fill
        c.font      = BODY_FONT
        c.border    = thin_border()
        c.alignment = CENTER if col == 1 else WRAP

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 8    # Sl No
ws.column_dimensions["B"].width = 48   # Description
ws.column_dimensions["C"].width = 16   # Detected Sprint
ws.column_dimensions["D"].width = 14   # Assigned To
ws.column_dimensions["E"].width = 22   # Type Of Defect
ws.column_dimensions["F"].width = 50   # Action Taken
ws.column_dimensions["G"].width = 20   # Action Taken Date
ws.column_dimensions["H"].width = 18   # Status
ws.column_dimensions["I"].width = 28   # Remarks

# ── Defect Data ────────────────────────────────────────────────────────────
# Columns: Sl No | Description | Detected Sprint | Assigned To |
#          Type Of Defect | Action Taken | Action Taken Date | Status | Remarks
defects = [
    (
        1,
        "syntax_validator.py was built during Milestone 1 but was never wired into "
        "the agent pipeline — validation function was never called before AI analysis.",
        "Sprint 1",
        "Developer",
        "Functional Defect",
        "Added 'validate_syntax' as a dedicated LangGraph node between "
        "detect_language and analysis agents. Integrated utils/syntax_validator.py "
        "into orchestrator.py. Added syntax_valid and syntax_error fields to AgentState.",
        "24-Jul-2026",
        "Closed",
        "Milestone 1 requirement missed in initial build"
    ),
    (
        2,
        "LangGraph conditional edge fan-out using lambda returning a list caused "
        "incorrect routing — parallel execution of Code and Security agents not working.",
        "Sprint 2",
        "Developer",
        "Integration Defect",
        "Replaced lambda with a proper route_after_syntax() routing function. "
        "Used explicit node list in add_conditional_edges(). Verified parallel "
        "fan-out to analyze_code and analyze_security nodes.",
        "24-Jul-2026",
        "Closed",
        "LangGraph API misuse — lambda routing not supported for fan-out"
    ),
    (
        3,
        "Model 'gemini-1.5-flash-latest' returns 404 NOT_FOUND error — "
        "Google deprecated the '-latest' suffix from all model aliases.",
        "Sprint 2",
        "Developer",
        "Environment Defect",
        "Removed deprecated '-latest' suffix from all model names in "
        "code_analysis.py, security_analysis.py, and orchestrator.py. "
        "Changed to 'gemini-1.5-flash' (stable version).",
        "24-Jul-2026",
        "Closed",
        "Google API deprecated all '-latest' model aliases"
    ),
    (
        4,
        "Embedding model 'models/text-embedding-004' returns 404 NOT_FOUND "
        "on the v1beta API endpoint used by langchain-google-genai.",
        "Sprint 2",
        "Developer",
        "Environment Defect",
        "Replaced Google GenerativeAI Embeddings with local FastEmbedEmbeddings "
        "(BAAI/bge-small-en-v1.5) via langchain-community. Added fastembed to "
        "requirements.txt. Deleted and re-indexed ChromaDB with new model.",
        "24-Jul-2026",
        "Closed",
        "Google v1beta API endpoint does not expose embedding models"
    ),
    (
        5,
        "pip install sentence-transformers fails with [WinError 206] — "
        "PyTorch dependency creates file paths exceeding Windows 260-character MAX_PATH limit.",
        "Sprint 2",
        "Developer",
        "Environment Defect",
        "Uninstalled torch (2.13.0) and removed sentence-transformers and "
        "langchain-huggingface from requirements. Replaced with fastembed "
        "which uses lightweight ONNX models with short file paths.",
        "24-Jul-2026",
        "Closed",
        "Windows MAX_PATH (260 chars) exceeded by PyTorch internal directory structure"
    ),
    (
        6,
        "Model 'gemini-2.0-flash' returns RESOURCE_EXHAUSTED (429) error — "
        "free tier quota for this model is near-zero RPM.",
        "Sprint 2",
        "Developer",
        "Environment Defect",
        "Switched LLM provider entirely to Groq API. Updated all three agent files "
        "to use ChatGroq with 'llama-3.3-70b-versatile' (analysis) and "
        "'llama-3.1-8b-instant' (language detection). Added GROQ_API_KEY to .env.",
        "24-Jul-2026",
        "Closed",
        "gemini-2.0-flash free tier has near-zero RPM; Groq offers 14,400 req/day free"
    ),
    (
        7,
        "Model 'gemini-1.5-flash' returns 404 NOT_FOUND — langchain-google-genai "
        "package uses v1beta API which does not support this model name.",
        "Sprint 2",
        "Developer",
        "Environment Defect",
        "Fully migrated all LLM calls from Google Gemini to Groq (langchain-groq). "
        "Removed langchain-google-genai from requirements.txt. All agents now use "
        "Groq's llama models which are stable and have no quota issues.",
        "24-Jul-2026",
        "Closed",
        "Old langchain-google-genai package incompatible with current Gemini model names"
    ),
]

# ── Write rows ─────────────────────────────────────────────────────────────
for idx, defect in enumerate(defects):
    row = idx + 2
    fill_row(ws, row, defect, alt=(idx % 2 == 1))

    # Colour the Status cell
    status_cell = ws.cell(row=row, column=8)
    if status_cell.value == "Closed":
        status_cell.fill = GREEN_FILL
        status_cell.font = GREEN_FONT
        status_cell.alignment = CENTER
    elif status_cell.value == "Open":
        status_cell.fill = RED_FILL
        status_cell.font = RED_FONT
        status_cell.alignment = CENTER

ws.freeze_panes = "A2"

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(FILE_PATH)
print("\n" + "=" * 55)
print("  Defect_Tracker Template_v0.1.xlsx — FILLED!")
print("=" * 55)
print(f"\n  Location : {FILE_PATH}")
print(f"  Defects written : {len(defects)}")
print("\n  Green  = Closed (bug fixed)")
print("\n  Open the file in WPS to review!")
print("\n  Next → run: python fill_agile_template.py")
